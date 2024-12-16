import os
import openpyxl
import re

def get_input_file(default_directory):
    """Prompt user to input the name of the Excel file."""
    while True:
        input_file = input("Enter the name of the input Excel file (with .xlsx extension): ")
        full_path = os.path.join(default_directory, input_file)  # Combine with default directory
        try:
            wb = openpyxl.load_workbook(full_path)  # Attempt to load the workbook
            return full_path  # Return full path
        except FileNotFoundError:
            print(f"File '{full_path}' not found. Please try again.")

def get_columns_to_clean():
    """Prompt user to input the column indices to clean."""
    while True:
        try:
            columns_input = input("Enter the column indices to clean (comma-separated, e.g., 1,2 for columns A and B): ")
            columns = [int(col.strip()) for col in columns_input.split(',')]  # Return as is (1-based index)
            if any(col < 1 for col in columns):
                print("Column index must be at least 1.")
                continue
            return columns  # Return list of column indices
        except ValueError:
            print("Invalid input. Please enter valid integers.")

def get_output_file_details(default_directory):
    """Prompt user for output file name and directory."""
    output_directory = input(f"Enter the output directory (or '0' to use '{default_directory}'): ")
    
    if output_directory == '0':
        output_directory = default_directory
    
    output_file_name = input("Enter the output file name (without extension): ")
    
    full_output_path = os.path.join(output_directory, f"{output_file_name}.xlsx")
    
    return full_output_path

def format_date(date_string):
    """Format date strings to 'XXXX年X月X日'."""
    cleaned_string = ''.join(date_string.split())
    
    # Pattern 1: YYYY-MM-DD
    match = re.match(r'(\d{4})-(\d{1,2})-(\d{1,2})', cleaned_string)
    if match:
        year = match.group(1)
        month = str(int(match.group(2)))  # Convert to int to remove leading zeros
        day = str(int(match.group(3)))    # Convert to int to remove leading zeros
        return f"{year}年{month}月{day}日"
    
    # Pattern 2: YYYY年MM月DD日
    match = re.match(r'(\d{4})年(\d{1,2})月(\d{1,2})日', cleaned_string)
    if match:
        year = match.group(1)
        month = str(int(match.group(2)))  # Convert to int to remove leading zeros
        day = str(int(match.group(3)))    # Convert to int to remove leading zeros
        return f"{year}年{month}月{day}日"
    
    # Pattern 3: YYYYMMDD
    match = re.match(r'(\d{4})(\d{2})(\d{2})', cleaned_string)
    if match:
        year = match.group(1)
        month = str(int(match.group(2)))  # Convert to int to remove leading zeros
        day = str(int(match.group(3)))    # Convert to int to remove leading zeros
        return f"{year}年{month}月{day}日"
    
    # If date format does not match any pattern, return None
    return None

def clean_dates(input_file, column_indices, output_file):
    """Clean and format dates in specified columns of an Excel file."""
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    
    changes_made = 0
    total_processed = 0

    for col in column_indices:
        for row in range(1, ws.max_row + 1):  # Start from row 1 (including header)
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str):  # Check if the cell contains text
                total_processed += 1
                formatted_date = format_date(cell.value.strip())
                if formatted_date:
                    if cell.value != formatted_date:  # Only count if the value actually changed
                        cell.value = formatted_date
                        changes_made += 1

    # Save the changes to a new Excel file
    wb.save(output_file)
    print(f"Processing complete:")
    print(f"Total cells processed: {total_processed}")
    print(f"Changes made: {changes_made}")
    print(f"Results saved to: {output_file}")

# Main execution flow
default_directory = os.getcwd()

input_excel_file = get_input_file(default_directory)  # Pass default directory here
columns_to_clean = get_columns_to_clean()
output_excel_file = get_output_file_details(default_directory)

clean_dates(input_excel_file, columns_to_clean, output_excel_file)
