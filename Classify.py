from openpyxl import load_workbook
from collections import Counter
import os

def save_classifications(classifications, filename="classifications.txt"):
    """Save classifications to a text file with UTF-8 encoding."""
    with open(filename, 'w', encoding='utf-8') as file:
        for item, category in classifications.items():
            file.write(f"{item}:{category}\n")
    print("Classifications saved to", filename)

def load_classifications(filename="classifications.txt"):
    """Load classifications from a text file."""
    if not os.path.exists(filename):
        return {}
    
    classifications = {}
    with open(filename, 'r', encoding='utf-8') as file:
        for line in file:
            item, category = line.strip().split(':')
            classifications[item] = category
    print("Classifications loaded from", filename)
    return classifications

def classify_products():
    """Classify products as Food, Drink, or Others based on user input from an Excel file."""
    classifications = load_classifications()  # Load existing classifications
    classification_map = {
        'f': 'Food',
        'd': 'Drink',
        'o': 'Others'
    }

    while True:
        print("\nSelect an option:")
        print("0 - Classify products")
        print("1 - View classifications")
        print("2 - Output classifications from Excel")
        print("3 - Count classifications")
        print("4 - Manage classifications")
        print("5 - Save classifications")
        print("6 - Exit")
        
        choice = input("Enter your choice (0, 1, 2, 3, 4, 5, or 6): ")
        
        if choice == '0':
            # Classify products
            file_name = input("Enter the name of the Excel file (with .xlsx extension): ")
            try:
                # Load the Excel workbook
                wb = load_workbook(file_name)
                sheet = wb.active  # Get the active sheet
                print("Available columns:", [sheet.cell(row=1, column=i).value for i in range(1, sheet.max_column + 1)])
                
                column_number = int(input("Enter the column number (starting from 1): "))
                
                # Extract unique items from the designated column
                if 1 <= column_number <= sheet.max_column:
                    items = set()
                    for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip header
                        item = sheet.cell(row=row, column=column_number).value
                        if item is not None:
                            items.add(item)
                    
                    # Ask user to classify each unique item
                    for item in items:
                        classification = input(f"Is '{item}' food (f), drink (d), or others (o)? ").strip().lower()
                        while classification not in classification_map.keys():
                            classification = input(f"Invalid input. Please enter {', '.join(classification_map.keys())}: ").strip().lower()
                        classifications[item] = classification_map[classification]
                    
                    print("Classifications saved.")
                else:
                    print("Invalid column number.")
            except FileNotFoundError:
                print(f"Error: The file '{file_name}' was not found.")
            except Exception as e:
                print(f"An error occurred: {e}")

        elif choice == '1':
            # View classifications
            if not classifications:
                print("No classifications available. Please classify products first.")
                continue
            
            for item, category in classifications.items():
                print(f"{item} is classified as {category}.")
        
        elif choice == '2':
            # Output classifications from Excel
            file_name = input("Enter the name of the Excel file (with .xlsx extension): ")
            try:
                wb = load_workbook(file_name)
                sheet = wb.active  # Get the active sheet
                print("Available columns:", [sheet.cell(row=1, column=i).value for i in range(1, sheet.max_column + 1)])
                
                column_number = int(input("Enter the column number (starting from 1): "))
                
                if 1 <= column_number <= sheet.max_column:
                    for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip header
                        item = sheet.cell(row=row, column=column_number).value
                        if item is not None:
                            category = classifications.get(item, "Unclassified")
                            print(f"{item} is classified as {category}.")
                else:
                    print("Invalid column number.")
            except FileNotFoundError:
                print(f"Error: The file '{file_name}' was not found.")
            except Exception as e:
                print(f"An error occurred: {e}")

        elif choice == '3':
            # Count classifications from an Excel file and designated column
            file_name = input("Enter the name of the Excel file (with .xlsx extension): ")
            try:
                wb = load_workbook(file_name)
                sheet = wb.active  # Get the active sheet
                print("Available columns:", [sheet.cell(row=1, column=i).value for i in range(1, sheet.max_column + 1)])
                
                column_number = int(input("Enter the column number (starting from 1): "))
                
                if 1 <= column_number <= sheet.max_column:
                    counts = Counter()
                    for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip header
                        item = sheet.cell(row=row, column=column_number).value
                        if item is not None:
                            category = classifications.get(item)
                            if category: 
                                counts[category] += 1
                    
                    print("\nClassification Counts:")
                    for category, count in counts.items():
                        print(f"{category}: {count}")
                else:
                    print("Invalid column number.")
            except FileNotFoundError:
                print(f"Error: The file '{file_name}' was not found.")
            except Exception as e:
                print(f"An error occurred: {e}")

        elif choice == '4':
            # Manage classifications (add or delete)
            action = input("Do you want to add a new classification (a) or delete an existing one (d)? ").strip().lower()
            
            if action == 'a':
                new_classification = input("Enter the name of the new classification: ").strip()
                abbreviation = input("Enter an abbreviation for this classification: ").strip().lower()
                
                if abbreviation in classification_map.keys():
                    print(f"Abbreviation '{abbreviation}' already exists. Please choose a different one.")
                else:
                    classification_map[abbreviation] = new_classification.capitalize()
                    print(f"New classification '{new_classification}' with abbreviation '{abbreviation}' added.")
            
            elif action == 'd':
                abbreviation_to_delete = input("Enter the abbreviation of the classification to delete: ").strip().lower()
                
                if abbreviation_to_delete in classification_map:
                    del classification_map[abbreviation_to_delete]
                    print(f"Classification with abbreviation '{abbreviation_to_delete}' has been deleted.")
                else:
                    print(f"No classification found for abbreviation '{abbreviation_to_delete}'.")
            
            else:
                print("Invalid action. Please enter 'a' to add or 'd' to delete.")

        elif choice == '5':
            # Save classifications to a text file
            save_classifications(classifications)

        elif choice == '6':
            print("Exiting the program.")
            break
        
        else:
            print("Invalid choice. Please enter 0, 1, 2, 3, 4, 5, or 6.")

# Run the program
if __name__ == "__main__":
    classify_products()
