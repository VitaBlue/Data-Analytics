import os
import openpyxl
from difflib import SequenceMatcher

# Predefined product list
PRODUCT_LIST = []

def get_input_file(default_directory):
    """Prompt user to input the name of the Excel file."""
    while True:
        input_file = input("Enter the name of the input Excel file (with .xlsx extension): ")
        full_path = os.path.join(default_directory, input_file)
        try:
            wb = openpyxl.load_workbook(full_path)
            return full_path
        except FileNotFoundError:
            print(f"File '{full_path}' not found. Please try again.")

def get_columns_to_clean():
    """Prompt user to input the column indices to clean."""
    while True:
        try:
            columns_input = input("Enter the column indices to clean (comma-separated, e.g., 1,2 for columns A and B): ")
            columns = [int(col.strip()) for col in columns_input.split(',')]
            if any(col < 1 for col in columns):
                print("Column index must be at least 1.")
                continue
            return columns
        except ValueError:
            print("Invalid input. Please enter valid integers.")

def get_output_file_details(default_directory):
    """Prompt user for output file name and directory."""
    output_directory = input(f"Enter the output directory (or '0' to use '{default_directory}'): ")
    
    if output_directory == '0':
        output_directory = default_directory
    
    output_file_name = input("Enter the output file name (without extension): ")
    
    full_output_path = os.path.join(output_directory, f"{output_file_name}.xlsx")
    
    return full_output_path

def similarity_ratio(a, b):
    """Calculate similarity ratio between two strings."""
    return SequenceMatcher(None, a, b).ratio()

def get_sorted_matches(input_text, products, n=5):
    """Get sorted matches with similarity ratios."""
    similarities = [(p, similarity_ratio(input_text, p)) for p in products]
    return sorted(similarities, key=lambda x: x[1], reverse=True)[:n]

def ask_for_confirmation(input_text, match, similarity):
    """Ask user to confirm if the matched product is correct."""
    print(f"\nOriginal product name: {input_text}")
    print(f"Suggested change to: {match}")
    print(f"Similarity: {similarity:.1%}")
    
    while True:
        response = input("Accept this change? (Y/N): ").strip().upper()
        if response in ['Y', 'N']:
            return response == 'Y'
        print("Please enter Y or N")

def select_from_list(input_text, matches):
    """Let user select from a list of potential matches."""
    print(f"\nOriginal product name: '{input_text}'")
    print("Please select the correct product name from the following options:")
    print("0. Keep original value (no change)")
    for i, (product, similarity) in enumerate(matches, 1):
        print(f"{i}. {product} (Similarity: {similarity:.1%})")
    
    while True:
        try:
            choice = int(input("\nPlease select the correct product number (0-{}): ".format(len(matches))))
            if 0 <= choice <= len(matches):
                if choice == 0:
                    print(f"Keeping original value: {input_text}")
                else:
                    print(f"Selected: {matches[choice-1][0]}")
                return matches[choice-1][0] if choice > 0 else None
        except ValueError:
            pass
        print(f"Please enter a number between 0 and {len(matches)}")

def find_closest_match(input_text, product_list, threshold=0.6):
    """Find the closest matching product name with user confirmation."""
    if not input_text or not isinstance(input_text, str):
        return None
    
    cleaned_text = input_text.strip()
    if not cleaned_text:
        return None
    
    # Get sorted matches
    matches = get_sorted_matches(cleaned_text, product_list)
    
    if not matches:
        return None
    
    best_match, similarity = matches[0]
    
    # If similarity is 85% or higher, return match directly
    if similarity >= 0.85:
        print(f"\nMatch found: {cleaned_text}")
        return best_match
    
    # If similarity is above threshold but below 85%, ask for confirmation
    if similarity >= threshold:
        if ask_for_confirmation(cleaned_text, best_match, similarity):
            return best_match
    
    # If similarity is low or user rejected the first match, show options list
    return select_from_list(cleaned_text, matches)

def get_product_list_file(default_directory):
    """Prompt user to input the name of the product list file."""
    while True:
        input_file = input("Enter the name of the product list file (with .txt extension): ")
        full_path = os.path.join(default_directory, input_file)
        try:
            with open(full_path, 'r', encoding='utf-8') as f:
                # Try to read the file to verify it exists and is readable
                products = [line.strip().strip('"\'') for line in f.readlines()]
                products = [p for p in products if p]
                if not products:
                    print("Warning: File is empty")
                return full_path, products
        except FileNotFoundError:
            print(f"File '{full_path}' not found. Please try again.")
        except Exception as e:
            print(f"Error reading file: {e}")

def load_product_list(filename):
    """Load product list from txt file."""
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            # 讀取每行，去除引號和空白
            products = [line.strip().strip('"\'') for line in f.readlines()]
            # 過濾掉空行
            products = [p for p in products if p]
        return products
    except FileNotFoundError:
        print(f"Warning: {filename} not found. Creating new file.")
        with open(filename, 'w', encoding='utf-8') as f:
            f.write("")
        return []

def save_product_list(new_products, filename):
    """Save product list to txt file while preserving existing products."""
    try:
        # 讀取現有的商品列表
        with open(filename, 'r', encoding='utf-8') as f:
            existing_products = [line.strip().strip('"\'') for line in f.readlines()]
            existing_products = [p for p in existing_products if p]
    except FileNotFoundError:
        existing_products = []

    # 合併現有商品和新商品，保持原有順序
    all_products = existing_products.copy()
    
    # 只添加不在現有列表中的新商品
    for product in new_products:
        if product not in existing_products:
            all_products.append(product)

    # 保存所有商品
    with open(filename, 'w', encoding='utf-8') as f:
        for product in all_products:
            f.write(f'"{product}"\n')

def handle_unmatched_items(unmatched_items, product_list_file):
    """Handle unmatched items and optionally add them to product list."""
    if not unmatched_items:
        return
    
    print("\nProcessing unmatched items:")
    new_products = set()
    
    for item in sorted(unmatched_items):
        while True:
            response = input(f'\nAdd "{item}" to product list? (Y/N): ').strip().upper()
            if response in ['Y', 'N']:
                if response == 'Y':
                    new_products.add(item)
                    print(f'Added "{item}" to product list')
                break
            print("Please enter Y or N")
    
    if new_products:
        save_product_list(new_products, product_list_file)
        print(f"\nUpdated product list saved to {product_list_file}")

def clean_product_names(input_file, column_indices, output_file, product_list_file, product_list):
    """Clean and standardize product names in specified columns of an Excel file."""
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    
    changes_made = 0
    total_processed = 0
    unmatched_items = set()
    # Store confirmed matches
    confirmed_matches = {}

    for col in column_indices:
        # Start from row 2 (skip header)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str):
                total_processed += 1
                original_value = cell.value.strip()
                print(f"\nProcessing row {row}, column {col}: {original_value}")
                
                # Check if we have a confirmed match
                if original_value in confirmed_matches:
                    matched_product = confirmed_matches[original_value]
                    if matched_product is not None and cell.value != matched_product:
                        cell.value = matched_product
                        changes_made += 1
                        print(f"Applied confirmed match: {matched_product}")
                    elif matched_product is None:
                        unmatched_items.add(original_value)
                        print("Applied previous decision: keeping original value")
                    continue
                
                # If no confirmed match exists, perform matching
                matched_product = find_closest_match(original_value, product_list)
                # Store the matching result (whether matched or not)
                confirmed_matches[original_value] = matched_product
                
                if matched_product:
                    if cell.value != matched_product:
                        cell.value = matched_product
                        changes_made += 1
                        print(f"Changed to: {matched_product}")
                else:
                    unmatched_items.add(original_value)
                    print("Keeping original value")

    # Save changes
    wb.save(output_file)
    
    # Print results
    print(f"\nProcessing complete:")
    print(f"Total cells processed: {total_processed}")
    print(f"Changes made: {changes_made}")
    
    if unmatched_items:
        print("\nUnmatched items:")
        for item in sorted(unmatched_items):
            print(f"- {item}")
        
        # Handle unmatched items
        handle_unmatched_items(unmatched_items, product_list_file)
    
    print(f"\nResults saved to: {output_file}")

def main():
    default_directory = os.getcwd()
    
    # Load product list from file
    product_list_file, product_list = get_product_list_file(default_directory)
    
    if not product_list:
        print("Error: Product list is empty. Please check your input file")
        return
    
    input_excel_file = get_input_file(default_directory)
    columns_to_clean = get_columns_to_clean()
    output_excel_file = get_output_file_details(default_directory)
    
    clean_product_names(input_excel_file, columns_to_clean, output_excel_file, product_list_file, product_list)

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nProgram terminated by user.")
    except Exception as e:
        print(f"\nAn error occurred: {str(e)}")