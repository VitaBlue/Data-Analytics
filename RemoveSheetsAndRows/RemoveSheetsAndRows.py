import openpyxl
import os

def delete_sheet(workbook, sheet_name):
    """Delete a specific sheet from the workbook."""
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
        print(f"Sheet '{sheet_name}' has been completely deleted.")
    else:
        print(f"Sheet '{sheet_name}' does not exist.")

def delete_rows(workbook, sheet_name, start_row, end_row):
    """Delete a range of rows from a given sheet in the workbook."""
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Create a list of merged cells to unmerge later
        merged_ranges = list(sheet.merged_cells.ranges)
        
        # Check for merged cells in the range to be deleted
        for merged_range in merged_ranges:
            if (merged_range.min_row <= end_row and merged_range.max_row >= start_row):
                # Unmerge cells if they are merged and overlap with the range to be deleted
                sheet.unmerge_cells(str(merged_range))
                print(f"Unmerged cells in range {merged_range} before deletion.")

        # Now delete the rows
        if 1 <= start_row <= end_row <= sheet.max_row:
            sheet.delete_rows(start_row, end_row - start_row + 1)
            print(f"Rows {start_row} to {end_row} have been deleted from '{sheet_name}'.")
        else:
            print(f"Row numbers {start_row} to {end_row} are out of range for sheet '{sheet_name}'.")
    else:
        print(f"Sheet '{sheet_name}' does not exist.")

def save_workbook(workbook, output_file):
    """Save the workbook to the specified output file."""
    workbook.save(output_file)
    print(f"Workbook saved as '{output_file}'.")

def display_menu():
    """Display the menu options."""
    print("\nMenu:")
    print("1. Delete a sheet")
    print("2. Delete rows")
    print("3. Save file and exit")
    print("Please select an option (1-3): ")

def main():
    # Load the input Excel file
    input_file = input("Enter the name of the input Excel file (with .xlsx extension): ")
    
    while not os.path.isfile(input_file):
        print(f"File '{input_file}' not found. Please try again.")
        input_file = input("Enter the name of the input Excel file (with .xlsx extension): ")

    workbook = openpyxl.load_workbook(input_file)

    while True:
        display_menu()
        choice = input()

        if choice == '1':
            # Delete a sheet
            sheet_name = input("Enter the name of the sheet to delete: ")
            delete_sheet(workbook, sheet_name)

        elif choice == '2':
            # Delete rows
            sheet_name = input("Enter the name of the sheet from which to delete rows: ")
            row_range = input("Enter the range of rows to delete (e.g., '2-5'): ")
            try:
                start_row, end_row = map(int, row_range.split('-'))
                delete_rows(workbook, sheet_name, start_row, end_row)
            except ValueError:
                print("Invalid row range format. Please enter in 'start-end' format.")

        elif choice == '3':
            # Save and exit
            output_file = input("Enter the name of the output file (with .xlsx extension): ")
            save_workbook(workbook, output_file)
            break  # Exit the loop

        else:
            print("Invalid option. Please choose between 1 and 3.")

if __name__ == "__main__":
    main()
