import os
import openpyxl

def get_input_file(default_directory):
    """Prompt user to input the name of the Excel file."""
    while True:
        input_file = input("Enter the name of the input Excel file (with .xlsx extension): ")
        full_path = os.path.join(default_directory, input_file)  # Combine with default directory
        try:
            wb = openpyxl.load_workbook(full_path)  # Attempt to load the workbook
            return full_path  # Return full path
        except FileNotFoundError:
            print(f"File '{full_path}' not found. Please try again.")

def get_columns_to_clean():
    """Prompt user to input the column indices to clean."""
    while True:
        try:
            columns_input = input("Enter the column indices to clean (comma-separated, e.g., 1,2 for columns A and B): ")
            columns = [int(col.strip()) for col in columns_input.split(',')]  # Return as is (1-based index)
            if any(col < 1 for col in columns):
                print("Column index must be at least 1.")
                continue
            return columns  # Return list of column indices
        except ValueError:
            print("Invalid input. Please enter valid integers.")

def get_output_file_details(default_directory):
    """Prompt user for output file name and directory."""
    output_directory = input(f"Enter the output directory (or '0' to use '{default_directory}'): ")
    
    if output_directory == '0':
        output_directory = default_directory
    
    output_file_name = input("Enter the output file name (without extension): ")
    
    full_output_path = os.path.join(output_directory, f"{output_file_name}.xlsx")
    
    return full_output_path

def remove_spaces_from_cells(input_file, column_indices, output_file):
    """Remove unnecessary spaces from cells in specified columns."""
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    for col in column_indices:
        for row in range(1, ws.max_row + 1):  # Start from row 1 (including header)
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str):  # Check if the cell contains text
                cleaned_value = ''.join(cell.value.split())  # Remove all spaces
                cell.value = cleaned_value  # Update cell value

    # Save the changes to a new Excel file
    wb.save(output_file)
    print(f"All unnecessary spaces have been removed and saved to {output_file}")

# Main execution flow
# Get default directory where the script is running
default_directory = os.getcwd()

# Get user input for file name and columns to clean
input_excel_file = get_input_file(default_directory)  # Pass default directory here
columns_to_clean = get_columns_to_clean()

# Get output file details from user
output_excel_file = get_output_file_details(default_directory)

# Remove spaces from specified columns and save to output file
remove_spaces_from_cells(input_excel_file, columns_to_clean, output_excel_file)
