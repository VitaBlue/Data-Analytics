import openpyxl
import os

def clean_numeric_columns(input_file, output_file, columns_to_clean):
    """Clean numeric data in specified columns of an Excel file."""
    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    # Iterate through each specified column
    for col in columns_to_clean:
        for row in range(2, ws.max_row + 1):  # Start from row 2 to skip header
            cell = ws.cell(row=row, column=col)  # Use col directly (1-based index)
            if isinstance(cell.value, str):  # Check if the cell contains text
                # Keep only numeric characters
                cleaned_value = ''.join(filter(str.isdigit, cell.value))
                # If cleaned_value is empty, set it to None
                cell.value = int(cleaned_value) if cleaned_value else None

    # Save the changes to a new Excel file
    wb.save(output_file)
    print(f"Cleaned numeric data and saved to {output_file}")

def get_input_file():
    """Prompt user to input the name of the Excel file."""
    while True:
        input_file = input("Enter the name of the input Excel file (with .xlsx extension): ")
        try:
            wb = openpyxl.load_workbook(input_file)  # Attempt to load the workbook
            return input_file
        except FileNotFoundError:
            print(f"File '{input_file}' not found. Please try again.")

def get_columns_to_clean():
    """Prompt user to input the column indices to clean."""
    while True:
        try:
            columns_input = input("Enter the column indices to clean (comma-separated, e.g., 1,2 for columns A and B): ")
            columns = [int(col.strip()) for col in columns_input.split(',')]  # No need to subtract 1 here
            if any(col < 1 for col in columns):  # Validate that all columns are at least 1
                print("Row or column values must be at least 1.")
                continue
            return columns
        except ValueError:
            print("Invalid input. Please enter valid integers.")

def get_output_file_details(default_directory):
    """Prompt user for output file name and directory."""
    output_directory = input(f"Enter the output directory (or '0' to use '{default_directory}'): ")
    
    if output_directory == '0':
        output_directory = default_directory
    
    output_file_name = input("Enter the output file name (without extension): ")
    
    full_output_path = os.path.join(output_directory, f"{output_file_name}.xlsx")
    
    return full_output_path

# Main execution flow
input_excel_file = get_input_file()
columns_to_clean = get_columns_to_clean()

default_directory = os.getcwd()
output_excel_file = get_output_file_details(default_directory)

clean_numeric_columns(input_excel_file, output_excel_file, columns_to_clean)
