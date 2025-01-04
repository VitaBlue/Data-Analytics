import openpyxl
from opencc import OpenCC
import os

def get_input_file(default_directory):
    """Prompt user to input the name of the Excel file."""
    while True:
        input_file = input("Enter the name of the input Excel file (with .xlsx extension): ")
        full_path = os.path.join(default_directory, input_file)  # Combine with default directory
        try:
            wb = openpyxl.load_workbook(full_path)  # Attempt to load the workbook
            return full_path  # Return full path
        except FileNotFoundError:
            print(f"File '{full_path}' not found. Please try again.")

def get_output_file_details(default_directory):
    """Prompt user for output file name and directory."""
    output_directory = input(f"Enter the output directory (or '0' to use '{default_directory}'): ")
    
    if output_directory == '0':
        output_directory = default_directory
    
    output_file_name = input("Enter the output file name (without extension): ")
    
    full_output_path = os.path.join(output_directory, f"{output_file_name}.xlsx")
    
    return full_output_path

def convert_simplified_to_traditional(input_file, output_file):
    """Convert Simplified Chinese text in an Excel file to Traditional Chinese."""
    cc = OpenCC('s2t')  # 's2t' stands for Simplified to Traditional

    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    # Iterate through all rows and columns in the sheet
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):  # Check if the cell contains text
                original_text = cell.value
                # Convert Simplified Chinese to Traditional Chinese
                converted_text = cc.convert(original_text)
                cell.value = converted_text  # Update the cell value

    # Save the changes to a new Excel file
    wb.save(output_file)
    print(f"Converted Simplified Chinese to Traditional Chinese and saved to {output_file}")

def main():
    """Main execution flow."""
    # Get default directory where the script is running
    default_directory = os.getcwd()

    # Get user input for file name
    input_excel_file = get_input_file(default_directory)

    # Get output file details from user
    output_excel_file = get_output_file_details(default_directory)

    # Call the conversion function with user-specified file name
    convert_simplified_to_traditional(input_excel_file, output_excel_file)

# Run the main function if this script is executed
if __name__ == "__main__":
    main()
