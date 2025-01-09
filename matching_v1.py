import os
import openpyxl
from difflib import SequenceMatcher

# Predefined product list
PRODUCT_LIST = [
    "酥香菠蘿油", "蒜香芝士麵包", "招牌海鹽卷", "日式吞拿魚麵包", "蒜香片",
    "薯仔煙肉拼日式碎蛋三文治", "香煎豬排沙律三文治", "吞拿魚拼日式碎蛋三文治",
    "醇香芝士夾火腿三文治", "日式碎蛋蟹柳三文治", "煙燻鴨胸沙律三文治",
    "紅豆生吐司", "椰蓉生吐司", "芝士火腿生吐司", "原味生吐司", "流淚生吐司",
    "榛子可可薄脆", "卡仕達泡芙", "美式曲奇", "紙杯蛋糕", "布丁蛋糕",
    "抹茶布丁蛋糕", "提拉米蘇", "巴斯克芝士蛋糕", "巴斯克開心果芝士蛋糕",
    "芝士火腿牛角包", "煙三文魚大蒜芝士牛角包", "原味牛角包",
    "紫菜肉鬆牛角窩夫", "牛角窩夫", "原味牛角窩夫", "焦糖餅", "⁠方包",
    "現有面團", "抹茶卷", "腸仔包", "菠蘿包面團", "⁠⁠貝果", "花椒餅乾",
    "牛油脆餅", "原味司康", "港式蛋撻", "蔓越莓司康", "脆脆薄片", "生吐司"
]

def get_input_file(default_directory):
    """Prompt user to input the name of the Excel file."""
    while True:
        input_file = input("Enter the name of the input Excel file (with .xlsx extension): ")
        full_path = os.path.join(default_directory, input_file)
        try:
            wb = openpyxl.load_workbook(full_path)
            return full_path
        except FileNotFoundError:
            print(f"File '{full_path}' not found. Please try again.")

def get_columns_to_clean():
    """Prompt user to input the column indices to clean."""
    while True:
        try:
            columns_input = input("Enter the column indices to clean (comma-separated, e.g., 1,2 for columns A and B): ")
            columns = [int(col.strip()) for col in columns_input.split(',')]
            if any(col < 1 for col in columns):
                print("Column index must be at least 1.")
                continue
            return columns
        except ValueError:
            print("Invalid input. Please enter valid integers.")

def get_output_file_details(default_directory):
    """Prompt user for output file name and directory."""
    output_directory = input(f"Enter the output directory (or '0' to use '{default_directory}'): ")
    
    if output_directory == '0':
        output_directory = default_directory
    
    output_file_name = input("Enter the output file name (without extension): ")
    
    full_output_path = os.path.join(output_directory, f"{output_file_name}.xlsx")
    
    return full_output_path

def similarity_ratio(a, b):
    """Calculate similarity ratio between two strings."""
    return SequenceMatcher(None, a, b).ratio()

def get_sorted_matches(input_text, products, n=5):
    """Get sorted matches with similarity ratios."""
    similarities = [(p, similarity_ratio(input_text, p)) for p in products]
    return sorted(similarities, key=lambda x: x[1], reverse=True)[:n]

def ask_for_confirmation(input_text, match, similarity):
    """Ask user to confirm if the matched product is correct."""
    print(f"\nOriginal product name: {input_text}")
    print(f"Suggested change to: {match}")
    print(f"Similarity: {similarity:.1%}")
    
    while True:
        response = input("Accept this change? (Y/N): ").strip().upper()
        if response in ['Y', 'N']:
            return response == 'Y'
        print("Please enter Y or N")

def select_from_list(input_text, matches):
    """Let user select from a list of potential matches."""
    print(f"\nOriginal product name: '{input_text}'")
    print("Please select the correct product name from the following options:")
    print("0. Keep original value (no change)")
    for i, (product, similarity) in enumerate(matches, 1):
        print(f"{i}. {product} (Similarity: {similarity:.1%})")
    
    while True:
        try:
            choice = int(input("\nPlease select the correct product number (0-{}): ".format(len(matches))))
            if 0 <= choice <= len(matches):
                if choice == 0:
                    print(f"Keeping original value: {input_text}")
                else:
                    print(f"Selected: {matches[choice-1][0]}")
                return matches[choice-1][0] if choice > 0 else None
        except ValueError:
            pass
        print(f"Please enter a number between 0 and {len(matches)}")

def find_closest_match(input_text, threshold=0.6):
    """Find the closest matching product name with user confirmation."""
    if not input_text or not isinstance(input_text, str):
        return None
    
    cleaned_text = input_text.strip()
    if not cleaned_text:
        return None
    
    # Get sorted matches
    matches = get_sorted_matches(cleaned_text, PRODUCT_LIST)
    
    if not matches:
        return None
    
    best_match, similarity = matches[0]
    
    # If similarity is 85% or higher, return match directly
    if similarity >= 0.85:
        print(f"\nMatch found: {cleaned_text}")
        return best_match
    
    # If similarity is above threshold but below 85%, ask for confirmation
    if similarity >= threshold:
        if ask_for_confirmation(cleaned_text, best_match, similarity):
            return best_match
    
    # If similarity is low or user rejected the first match, show options list
    return select_from_list(cleaned_text, matches)

def clean_product_names(input_file, column_indices, output_file):
    """Clean and standardize product names in specified columns of an Excel file."""
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    
    changes_made = 0
    total_processed = 0
    unmatched_items = set()
    # Store confirmed matches
    confirmed_matches = {}

    for col in column_indices:
        # Start from row 2 (skip header)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str):
                total_processed += 1
                original_value = cell.value.strip()
                print(f"\nProcessing row {row}, column {col}: {original_value}")
                
                # Check if we have a confirmed match
                if original_value in confirmed_matches:
                    matched_product = confirmed_matches[original_value]
                    if matched_product is not None and cell.value != matched_product:
                        cell.value = matched_product
                        changes_made += 1
                        print(f"Applied confirmed match: {matched_product}")
                    elif matched_product is None:
                        unmatched_items.add(original_value)
                        print("Applied previous decision: keeping original value")
                    continue
                
                # If no confirmed match exists, perform matching
                matched_product = find_closest_match(original_value)
                # Store the matching result (whether matched or not)
                confirmed_matches[original_value] = matched_product
                
                if matched_product:
                    if cell.value != matched_product:
                        cell.value = matched_product
                        changes_made += 1
                        print(f"Changed to: {matched_product}")
                else:
                    unmatched_items.add(original_value)
                    print("Keeping original value")

    # Save changes
    wb.save(output_file)
    
    # Print results
    print(f"\nProcessing complete:")
    print(f"Total cells processed: {total_processed}")
    print(f"Changes made: {changes_made}")
    
    if unmatched_items:
        print("\nUnmatched items:")
        for item in sorted(unmatched_items):
            print(f"- {item}")
    
    print(f"\nResults saved to: {output_file}")

def main():
    default_directory = os.getcwd()
    
    input_excel_file = get_input_file(default_directory)
    columns_to_clean = get_columns_to_clean()
    output_excel_file = get_output_file_details(default_directory)
    
    clean_product_names(input_excel_file, columns_to_clean, output_excel_file)

if __name__ == "__main__":
    main()