import os
import openpyxl
import re

def get_input_file(default_directory):
    """Prompt user to input the name of the Excel file."""
    while True:
        input_file = input("Enter the name of the input Excel file (with .xlsx extension): ")
        full_path = os.path.join(default_directory, input_file)
        try:
            wb = openpyxl.load_workbook(full_path)
            return full_path
        except FileNotFoundError:
            print(f"File '{full_path}' not found. Please try again.")

def get_columns_to_clean():
    """Prompt user to input the column indices to clean."""
    while True:
        try:
            columns_input = input("Enter the column indices to clean (comma-separated, e.g., 1,2 for columns A and B): ")
            columns = [int(col.strip()) for col in columns_input.split(',')]
            if any(col < 1 for col in columns):
                print("Column index must be at least 1.")
                continue
            return columns
        except ValueError:
            print("Invalid input. Please enter valid integers.")

def get_output_file_details(default_directory):
    """Prompt user for output file name and directory."""
    output_directory = input(f"Enter the output directory (or '0' to use '{default_directory}'): ")
    
    if output_directory == '0':
        output_directory = default_directory
    
    output_file_name = input("Enter the output file name (without extension): ")
    
    full_output_path = os.path.join(output_directory, f"{output_file_name}.xlsx")
    
    return full_output_path

def format_time(time_string):
    """Format time strings to 'HH:MM' in 24-hour format."""
    if not isinstance(time_string, str):
        return "N/A"
    
    # Remove all whitespace and check if empty
    cleaned_string = ''.join(time_string.split())
    if not cleaned_string:  # Handle empty strings or strings with only spaces
        return "N/A"
    
    # Pattern 1: HH:MM or H:MM (24-hour format)
    match = re.match(r'(\d{1,2}):(\d{1,2})', cleaned_string)
    if match:
        hour = int(match.group(1))
        minute = int(match.group(2))
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            return f"{hour:02d}:{minute:02d}"
    
    # Pattern 2: HHMM (24-hour format)
    match = re.match(r'(\d{2})(\d{2})', cleaned_string)
    if match:
        hour = int(match.group(1))
        minute = int(match.group(2))
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            return f"{hour:02d}:{minute:02d}"
    
    # Pattern 3: H:MM or HH:M (handle single digits and different separators)
    match = re.match(r'(\d{1,2})[:.。](\d{1,2})', cleaned_string)
    if match:
        hour = int(match.group(1))
        minute = int(match.group(2))
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            return f"{hour:02d}:{minute:02d}"
    
    # If time format does not match any pattern or is invalid, return 'N/A'
    return "N/A"

def clean_times(input_file, column_indices, output_file):
    """Clean and format times in specified columns of an Excel file."""
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    
    changes_made = 0
    total_processed = 0

    for col in column_indices:
        for row in range(2, ws.max_row + 1):  # Start from row 2 to skip header
            cell = ws.cell(row=row, column=col)
            total_processed += 1
            
            # Handle empty cells, None values, or cells with only spaces
            if cell.value is None or (isinstance(cell.value, str) and not cell.value.strip()):
                if cell.value != "N/A":  # Only update if it's not already "N/A"
                    cell.value = "N/A"
                    changes_made += 1
                continue
            
            formatted_time = format_time(str(cell.value).strip())
            if cell.value != formatted_time:
                cell.value = formatted_time
                changes_made += 1

    # Save the changes to a new Excel file
    wb.save(output_file)
    print(f"Processing complete:")
    print(f"Total cells processed: {total_processed}")
    print(f"Changes made: {changes_made}")
    print(f"Results saved to: {output_file}")

# Main execution flow
default_directory = os.getcwd()

input_excel_file = get_input_file(default_directory)
columns_to_clean = get_columns_to_clean()
output_excel_file = get_output_file_details(default_directory)

clean_times(input_excel_file, columns_to_clean, output_excel_file)