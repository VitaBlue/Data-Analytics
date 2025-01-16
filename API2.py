from openpyxl import load_workbook
from collections import Counter
import os
import logging
from groq import Groq

# Set up logging
logging.basicConfig(level=logging.INFO)

# Load the API key from environment variables
api_key = os.getenv("GROQ_API_KEY")
if api_key is None:
    raise ValueError("API key not found. Please set the GROQ_API_KEY environment variable.")

# Initialize the Groq client with the API key
client = Groq(api_key=api_key)

def save_classifications(classifications, filename="classifications.txt"):
    """Save classifications to a text file with UTF-8 encoding."""
    with open(filename, 'w', encoding='utf-8') as file:
        for item, category in classifications.items():
            file.write(f"{item}:{category}\n")
    print("Classifications saved to", filename)

def load_classifications(filename="classifications.txt"):
    """Load classifications from a text file."""
    if not os.path.exists(filename):
        return {}
    
    classifications = {}
    with open(filename, 'r', encoding='utf-8') as file:
        for line in file:
            stripped_line = line.strip()
            if stripped_line:  # Only process non-empty lines
                parts = stripped_line.split(':')
                if len(parts) == 2:  # Ensure there are exactly two parts
                    item, category = parts
                    classifications[item] = category
                else:
                    print(f"Warning: Skipping malformed line: '{stripped_line}'")
    print("Classifications loaded from", filename)
    return classifications

def get_ai_response(product_name):
    """Function to get a response from the AI model based on product name."""
    model = "llama-3.3-70b-versatile"  # Ensure this is a valid model name
    messages = [
        {"role": "system", "content": "You are a helpful assistant."},
        {"role": "user", "content": f"Please classify '{product_name}' as food, beverage, or other. Only respond with 'food', 'beverage', or 'other'."}
    ]

    try:
        response = client.chat.completions.create(
            model=model,
            messages=messages,
            max_tokens=10  # Limit tokens to avoid excessive output
        )

        # Log response for debugging (optional)
        logging.debug("API Response: %s", response)

        # Extract and return the response message correctly
        if hasattr(response, 'choices') and len(response.choices) > 0:
            choice = response.choices[0]
            if hasattr(choice, 'message'):
                return choice.message.content.strip().lower()  # Return content in lowercase
            else:
                return "Invalid response."
        else:
            return "No choices in response."

    except Exception as e:
        logging.error("Error while calling Groq API: %s", e)
        return 'other'  # Default to 'other' on error

def classify_products():
    """Classify products as Food, Drink, or Others based on user input from an Excel file."""
    classifications = load_classifications()  # Load existing classifications

    while True:
        print("\nSelect an option:")
        print("0 - Classify products")
        print("1 - View classifications")
        print("2 - Output classifications from Excel")
        print("3 - Count classifications")
        print("4 - Manage classifications")
        print("5 - Save classifications")
        print("6 - Clear all classifications")
        print("7 - Exit")
        
        choice = input("Enter your choice (0, 1, 2, 3, 4, 5, 6, or 7): ")
        
        if choice == '0':
            # Classify products
            file_name = input("Enter the name of the Excel file (with .xlsx extension): ")
            try:
                # Load the Excel workbook
                wb = load_workbook(file_name)
                sheet = wb.active  # Get the active sheet
                print("Available columns:", [sheet.cell(row=1, column=i).value for i in range(1, sheet.max_column + 1)])
                
                column_number = int(input("Enter the column number (starting from 1): "))
                
                # Extract unique items from the designated column
                if 1 <= column_number <= sheet.max_column:
                    items = set()
                    for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip header
                        item = sheet.cell(row=row, column=column_number).value
                        if item is not None:
                            items.add(item)
                    
                    # Limit to first 20 unique items and classify them
                    classified_count = 0
                    
                    for item in items:
                        if classified_count >= 20:   # Stop after classifying 20 products
                            break
                        if item in classifications:
                            print(f"'{item}' is already classified as {classifications[item]}. Skipping.")
                        else:
                            category_result = get_ai_response(item)  # Call AI function to classify using Groq API

                            # Validate and ensure only allowed responses are stored
                            valid_categories = ['food', 'beverage', 'other']
                            if category_result in valid_categories:
                                classifications[item] = category_result.capitalize()  # Store classification
                                classified_count += 1   # Increment count of classified products
                    
                    print(f"Classified {classified_count} new products.")
                else:
                    print("Invalid column number.")
            except FileNotFoundError:
                print(f"Error: The file '{file_name}' was not found.")
            except Exception as e:
                print(f"An error occurred: {e}")

        elif choice == '1':
            # View classifications
            if not classifications:
                print("No classifications available. Please classify products first.")
                continue
            
            for item, category in classifications.items():
                print(f"{item} is classified as {category}.")
        
        elif choice == '2':
            # Output classifications from Excel
            file_name = input("Enter the name of the Excel file (with .xlsx extension): ")
            try:
                wb = load_workbook(file_name)
                sheet = wb.active  # Get the active sheet
                print("Available columns:", [sheet.cell(row=1, column=i).value for i in range(1, sheet.max_column + 1)])
                
                column_number = int(input("Enter the column number (starting from 1): "))
                
                if 1 <= column_number <= sheet.max_column:
                    for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip header
                        item = sheet.cell(row=row, column=column_number).value
                        if item is not None:
                            category = classifications.get(item, "Unclassified")
                            print(f"{item} is classified as {category}.")
                else:
                    print("Invalid column number.")
            except FileNotFoundError:
                print(f"Error: The file '{file_name}' was not found.")
            except Exception as e:
                print(f"An error occurred: {e}")

        elif choice == '3':
            # Count classifications from an Excel file and designated column
            file_name = input("Enter the name of the Excel file (with .xlsx extension): ")
            try:
                wb = load_workbook(file_name)
                sheet = wb.active  # Get the active sheet
                print("Available columns:", [sheet.cell(row=1, column=i).value for i in range(1, sheet.max_column + 1)])
                
                column_number = int(input("Enter the column number (starting from 1): "))
                
                if 1 <= column_number <= sheet.max_column:
                    counts = Counter()
                    for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip header
                        item = sheet.cell(row=row, column=column_number).value
                        if item is not None:
                            category = classifications.get(item)
                            if category: 
                                counts[category] += 1
                    
                    print("\nClassification Counts:")
                    for category, count in counts.items():
                        print(f"{category}: {count}")
                else:
                    print("Invalid column number.")
            except FileNotFoundError:
                print(f"Error: The file '{file_name}' was not found.")
            except Exception as e:
                print(f"An error occurred: {e}")

        elif choice == '4':
            # Manage classifications (add or delete)
            action = input("Do you want to add a new classification (a) or delete an existing one (d)? ").strip().lower()
            
            if action == 'a':
                new_classification = input("Enter the name of the new classification: ").strip()
                abbreviation = input("Enter an abbreviation for this classification: ").strip().lower()
                
                if abbreviation in classification_map.keys():
                    print(f"Abbreviation '{abbreviation}' already exists. Please choose a different one.")
                else:
                    classification_map[abbreviation] = new_classification.capitalize()
                    print(f"New classification '{new_classification}' with abbreviation '{abbreviation}' added.")
            
            elif action == 'd':
                abbreviation_to_delete = input("Enter the abbreviation of the classification to delete: ").strip().lower()
                
                if abbreviation_to_delete in classification_map:
                    del classification_map[abbreviation_to_delete]
                    print(f"Classification with abbreviation '{abbreviation_to_delete}' has been deleted.")
                else:
                    print(f"No classification found for abbreviation '{abbreviation_to_delete}'.")
            
            else:
                print("Invalid action. Please enter 'a' to add or 'd' to delete.")

        elif choice == '5':
            # Save classifications to a text file
            save_classifications(classifications)

        elif choice == '6':
            # Clear all classifications
            confirm_clear = input("Are you sure you want to clear all classifications? (yes/no): ").strip().lower()
            if confirm_clear == 'yes':
                classifications.clear()
                save_classifications(classifications)  # Optionally save empty state
                print("All classifications have been cleared.")
            else:
                print("Clear operation canceled.")

        elif choice == '7':
            print("Exiting the program.")
            break
        
        else:
            print("Invalid choice. Please enter 0, 1, 2, 3, 4, 5, 6, or 7.")

# Run the program
if __name__ == "__main__":
    classify_products()
