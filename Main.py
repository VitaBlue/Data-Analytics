import os
import openpyxl
import re
from opencc import OpenCC

def get_input_file(default_directory):
    """Prompt user to input the name of the Excel file."""
    while True:
        input_file = input("Enter the name of the input Excel file (with .xlsx extension): ")
        full_path = os.path.join(default_directory, input_file)
        try:
            wb = openpyxl.load_workbook(full_path)
            return full_path
        except FileNotFoundError:
            print(f"File '{full_path}' not found. Please try again.")

def get_output_file_details(default_directory):
    """Prompt user for output file name and directory."""
    output_directory = input(f"Enter the output directory (or '0' to use '{default_directory}'): ")
    
    if output_directory == '0':
        output_directory = default_directory
    
    output_file_name = input("Enter the output file name (without extension): ")
    
    full_output_path = os.path.join(output_directory, f"{output_file_name}.xlsx")
    
    return full_output_path

def get_column_headers(file_path):
    """Get the headers (first row) of the Excel file."""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    max_col = ws.max_column
    headers = []
    
    for col in range(1, max_col + 1):
        cell_value = ws.cell(row=1, column=col).value
        headers.append(cell_value if cell_value is not None else "")
        
    wb.close()
    return headers

def convert_simplified_to_traditional(input_file, output_file):
    """Convert Simplified Chinese text in an Excel file to Traditional Chinese."""
    cc = OpenCC('s2t')
    wb = openpyxl.load_workbook(input_file)
    
    special_changes = 0
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    # 先進行一般的繁簡轉換
                    converted_text = cc.convert(cell.value)
                    
                    # 特殊處理：將「餅幹」轉換為「餅乾」
                    if '餅幹' in converted_text:
                        original = converted_text
                        converted_text = converted_text.replace('餅幹', '餅乾')
                        if original != converted_text:
                            special_changes += 1
                    
                    cell.value = converted_text
    
    wb.save(output_file)
    if special_changes > 0:
        print(f"Chinese conversion completed. Special conversion ('餅幹' -> '餅乾'): {special_changes} changes")
    else:
        print("Chinese conversion completed.")
    return output_file

def process_text_column(ws, col):
    """Process text column - Remove spaces"""
    changes = 0
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell.value, str):
            cleaned_value = ''.join(cell.value.split())
            if cell.value != cleaned_value:
                cell.value = cleaned_value
                changes += 1
    return changes

def process_date_column(ws, col):
    """Process date column - Remove spaces and format date"""
    changes = 0
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell.value, str):
            # First remove spaces
            cleaned_value = ''.join(cell.value.split())
            
            # Then format date
            date_patterns = [
                r'(\d{4})-(\d{1,2})-(\d{1,2})',
                r'(\d{4})年(\d{1,2})月(\d{1,2})日',
                r'(\d{4})(\d{2})(\d{2})'
            ]
            
            formatted_date = None
            for pattern in date_patterns:
                match = re.match(pattern, cleaned_value)
                if match:
                    year = match.group(1)
                    month = str(int(match.group(2)))
                    day = str(int(match.group(3)))
                    formatted_date = f"{year}年{month}月{day}日"
                    break
            
            if formatted_date and cell.value != formatted_date:
                cell.value = formatted_date
                changes += 1
    return changes

def process_time_column(ws, col):
    """Process time column - Format time or replace with N/A"""
    changes = 0
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col)
        original_value = cell.value
        
        # Initialize cleaned_value
        cleaned_value = ""
        
        # Check if cell value is string
        if isinstance(cell.value, str):
            # Try to match time patterns first
            time_str = cell.value.strip()
            # Pattern for H:MM or HH:MM
            match = re.match(r'^(\d{1,2}):(\d{2})$', time_str)
            if match:
                hours = int(match.group(1))
                minutes = int(match.group(2))
                if 0 <= hours <= 23 and 0 <= minutes <= 59:
                    new_value = f"{hours:02d}:{minutes:02d}"
                    if cell.value != new_value:
                        cell.value = new_value
                        changes += 1
                    continue
            
            # If no match, try to extract numbers
            cleaned_value = ''.join(filter(str.isdigit, time_str))
        
        # Format time if we have valid numeric data
        new_value = "N/A"
        if cleaned_value:
            if len(cleaned_value) >= 4:
                hours = int(cleaned_value[:2])
                minutes = int(cleaned_value[2:4])
                if 0 <= hours <= 23 and 0 <= minutes <= 59:
                    new_value = f"{hours:02d}:{minutes:02d}"
            elif len(cleaned_value) == 3:
                hours = int(cleaned_value[0])
                minutes = int(cleaned_value[1:])
                if 0 <= hours <= 23 and 0 <= minutes <= 59:
                    new_value = f"{hours:02d}:{minutes:02d}"
        
        if original_value != new_value:
            cell.value = new_value
            changes += 1
            
    return changes

def process_number_column(ws, col):
    """Process number column - Remove text and spaces"""
    changes = 0
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell.value, str):
            # Keep only numeric characters
            cleaned_value = ''.join(filter(str.isdigit, cell.value))
            new_value = int(cleaned_value) if cleaned_value else None
            if cell.value != str(new_value):
                cell.value = new_value
                changes += 1
    return changes

def main():
    default_directory = os.getcwd()
    
    # Initialize counters for each type of cleaning
    total_text_changes = 0
    total_date_changes = 0
    total_time_changes = 0
    total_number_changes = 0
    
    # Step 1: Get input file and convert Chinese characters
    input_file = get_input_file(default_directory)
    temp_output = os.path.join(default_directory, "temp_converted.xlsx")
    converted_file = convert_simplified_to_traditional(input_file, temp_output)
    
    # Step 2: Get column headers and process based on type
    headers = get_column_headers(converted_file)
    print("\nColumn headers:")
    for i, header in enumerate(headers, 1):
        print(f"{i}. {header}")
    
    # Create new workbook for final output
    wb = openpyxl.load_workbook(converted_file)
    ws = wb.active
    
    # Process each column
    for col_num in range(1, len(headers) + 1):
        while True:
            print(f"\nFor column {col_num} ({headers[col_num-1]}), choose data type:")
            print("1. Text")
            print("2. Date")
            print("3. Time")
            print("4. Number")
            print("5. None")
            choice = input("Enter choice (1-5): ")
            
            if choice in ['1', '2', '3', '4', '5']:
                if choice == '1':  # Text
                    changes = process_text_column(ws, col_num)
                    print(f"Text cleaning: {changes} cells modified")
                    total_text_changes += changes
                elif choice == '2':  # Date
                    changes = process_date_column(ws, col_num)
                    print(f"Date cleaning: {changes} cells modified")
                    total_date_changes += changes
                elif choice == '3':  # Time
                    changes = process_time_column(ws, col_num)
                    print(f"Time cleaning: {changes} cells modified")
                    total_time_changes += changes
                elif choice == '4':  # Number
                    changes1 = process_number_column(ws, col_num)
                    changes2 = process_text_column(ws, col_num)
                    print(f"Number cleaning: {changes1} cells modified")
                    print(f"Additional text cleaning: {changes2} cells modified")
                    total_number_changes += changes1
                    total_text_changes += changes2
                # Choice 5 (None) does nothing
                break
            else:
                print("Invalid choice. Please enter a number between 1 and 5.")
    
    # Print total changes before saving
    print("\nTotal changes made:")
    print(f"Text cleaning: {total_text_changes} cells modified")
    print(f"Date cleaning: {total_date_changes} cells modified")
    print(f"Time cleaning: {total_time_changes} cells modified")
    print(f"Number cleaning: {total_number_changes} cells modified")
    print(f"Total modifications: {total_text_changes + total_date_changes + total_time_changes + total_number_changes} cells")
    
    # Save final output
    final_output = get_output_file_details(default_directory)
    wb.save(final_output)
    
    # Clean up temporary file
    try:
        os.remove(temp_output)
    except:
        pass
    
    print(f"\nProcessing complete. Final output saved to: {final_output}")

if __name__ == "__main__":
    main()
