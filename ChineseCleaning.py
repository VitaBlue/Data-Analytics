import openpyxl
from opencc import OpenCC
import os

def convert_simplified_to_traditional(input_file, output_file):
    """Convert Simplified Chinese text in an Excel file to Traditional Chinese."""
    # Initialize OpenCC for conversion from Simplified to Traditional Chinese
    cc = OpenCC('s2t')  # 's2t' stands for Simplified to Traditional

    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    # Iterate through all rows and columns in the sheet
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):  # Check if the cell contains text
                original_text = cell.value
                # Convert Simplified Chinese to Traditional Chinese
                converted_text = cc.convert(original_text)
                cell.value = converted_text  # Update the cell value

    # Save the changes to a new Excel file
    wb.save(output_file)
    print(f"Converted Simplified Chinese to Traditional Chinese and saved to {output_file}")

def get_input_file():
    """Prompt user to input the name of the Excel file."""
    while True:
        input_file = input("Enter the name of the input Excel file (with .xlsx extension): ")
        try:
            wb = openpyxl.load_workbook(input_file)  # Attempt to load the workbook
            return input_file
        except FileNotFoundError:
            print(f"File '{input_file}' not found. Please try again.")

def get_output_file_details(default_directory):
    """Prompt user for output file name and directory."""
    # Ask for output directory
    output_directory = input(f"Enter the output directory (or '0' to use '{default_directory}'): ")
    
    # If user inputs '0', maintain original directory
    if output_directory == '0':
        output_directory = default_directory
    
    # Ask for output file name
    output_file_name = input("Enter the output file name (without extension): ")
    
    # Combine directory and file name to create full path
    full_output_path = os.path.join(output_directory, f"{output_file_name}.xlsx")
    
    return full_output_path

# Main execution flow
input_excel_file = get_input_file()

# Get default directory where the script is running
default_directory = os.getcwd()

# Get output file details from user
output_excel_file = get_output_file_details(default_directory)

# Call the conversion function with user-specified file name
convert_simplified_to_traditional(input_excel_file, output_excel_file)